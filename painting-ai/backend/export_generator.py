"""
Export Generator - Create Excel and PDF outputs
"""

from typing import Dict, List
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from datetime import datetime, timedelta


class ExportGenerator:
    """Generate Excel and PDF exports for painting takeoffs"""

    def generate_excel(self, project: Dict, rooms: List[Dict], output_path: Path):
        """Generate detailed Excel takeoff"""

        workbook = openpyxl.Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create Summary sheet
        self._create_summary_sheet(workbook, project, rooms)

        # Create Detailed Takeoff sheet
        self._create_takeoff_sheet(workbook, project, rooms)

        # Create Room-by-Room sheet
        self._create_room_breakdown_sheet(workbook, project, rooms)

        # Save workbook
        workbook.save(output_path)

    def _create_summary_sheet(self, workbook, project: Dict, rooms: List[Dict]):
        """Create professional project summary sheet with branding"""

        sheet = workbook.create_sheet("Summary")

        # Set column widths for better appearance
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 20

        # Company Branding Section
        sheet["A1"] = "PAINTING.AI"
        sheet["A1"].font = Font(size=20, bold=True, color="1F4E78")
        sheet["A2"] = "Professional Painting Takeoff & Estimation"
        sheet["A2"].font = Font(size=10, italic=True, color="666666")

        # Header with background
        sheet["A4"] = "PAINTING PROPOSAL & ESTIMATE"
        sheet["A4"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A4"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        sheet.merge_cells('A4:C4')
        sheet["A4"].alignment = Alignment(horizontal='center')

        # Project Information Section
        row = 6
        info_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        sheet[f"A{row}"] = "PROJECT INFORMATION"
        sheet[f"A{row}"].font = Font(size=12, bold=True)
        sheet[f"A{row}"].fill = info_header_fill
        sheet.merge_cells(f'A{row}:C{row}')

        row += 1
        project_info = [
            ("Project Name:", project.get("name", ""), ""),
            ("Customer:", project.get("customer", ""), ""),
            ("Address:", project.get("address", "N/A"), ""),
            ("Project Type:", project.get("project_type", "Commercial").title(), ""),
            ("Date Prepared:", datetime.now().strftime("%B %d, %Y"), ""),
            ("Estimate Valid Until:", (datetime.now() + timedelta(days=30)).strftime("%B %d, %Y"), ""),
        ]

        for label, value, extra in project_info:
            sheet[f"A{row}"] = label
            sheet[f"B{row}"] = value
            sheet[f"C{row}"] = extra
            sheet[f"A{row}"].font = Font(bold=True)
            row += 1

        # Summary Totals Section
        row += 1
        sheet[f"A{row}"] = "PROJECT SUMMARY"
        sheet[f"A{row}"].font = Font(size=12, bold=True)
        sheet[f"A{row}"].fill = info_header_fill
        sheet.merge_cells(f'A{row}:C{row}')

        row += 1
        # Create bordered table for totals
        totals = [
            ("Total Rooms:", project.get("total_rooms", 0), "rooms"),
            ("Total Paintable Area:", f"{project.get('total_sqft', 0):,.0f}", "sqft"),
            ("Total Paint Required:", f"{project.get('total_gallons', 0):,.0f}", "gallons"),
            ("Total Labor Hours:", f"{project.get('total_labor_hours', 0):,.1f}", "hours"),
            ("Material Cost:", f"${project.get('estimated_cost', 0) * 0.35:,.2f}", ""),
            ("Labor Cost:", f"${project.get('estimated_cost', 0) * 0.65:,.2f}", ""),
            ("", "", ""),  # Spacer
            ("TOTAL ESTIMATE:", f"${project.get('estimated_cost', 0):,.2f}", ""),
        ]

        start_row = row
        for label, value, unit in totals:
            sheet[f"A{row}"] = label
            sheet[f"B{row}"] = value
            sheet[f"A{row}"].font = Font(bold=True)
            row += 1

        # Room list
        row += 2
        sheet[f"A{row}"] = "ROOMS"
        sheet[f"A{row}"].font = Font(size=14, bold=True)

        row += 1
        headers = ["Room Name", "Area (sqft)", "Walls", "Ceiling", "Trim"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

        row += 1
        for room in rooms:
            surfaces = room.get("surfaces", {})

            sheet.cell(row=row, column=1).value = room.get("name", "")
            sheet.cell(row=row, column=2).value = round(room.get("total_area", 0), 1)
            sheet.cell(row=row, column=3).value = round(surfaces.get("walls", {}).get("area", 0), 1)
            sheet.cell(row=row, column=4).value = round(surfaces.get("ceiling", {}).get("area", 0), 1)
            sheet.cell(row=row, column=5).value = round(surfaces.get("trim", {}).get("area", 0), 1)

            row += 1

        # Auto-size columns
        for col in range(1, 6):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        # Terms & Conditions Section
        row += 3
        sheet[f"A{row}"] = "TERMS & CONDITIONS"
        sheet[f"A{row}"].font = Font(size=12, bold=True)
        sheet[f"A{row}"].fill = info_header_fill
        sheet.merge_cells(f'A{row}:E{row}')

        row += 1
        terms = [
            "• This estimate is valid for 30 days from the date of preparation",
            "• Price includes all materials, labor, and supplies unless otherwise noted",
            "• Assumes normal working conditions and access to all areas",
            "• Additional charges may apply for extensive surface preparation, wallpaper removal, or lead paint",
            "• Payment terms: 50% deposit, balance upon completion",
            "• Warranty: 2 years on workmanship, manufacturer warranty on materials",
            "• Change orders must be approved in writing and will be billed separately",
            "• Final price subject to site inspection and condition assessment"
        ]

        for term in terms:
            sheet[f"A{row}"] = term
            sheet[f"A{row}"].font = Font(size=9)
            sheet.merge_cells(f'A{row}:E{row}')
            sheet[f"A{row}"].alignment = Alignment(wrap_text=True, vertical='top')
            sheet.row_dimensions[row].height = 20
            row += 1

        # Signature Block
        row += 2
        sheet[f"A{row}"] = "Prepared by:"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"C{row}"] = "Customer Acceptance:"
        sheet[f"C{row}"].font = Font(bold=True)

        row += 1
        sheet[f"A{row}"] = "____________________________"
        sheet[f"C{row}"] = "____________________________"

        row += 1
        sheet[f"A{row}"] = "Signature"
        sheet[f"A{row}"].font = Font(size=9, italic=True)
        sheet[f"C{row}"] = "Signature"
        sheet[f"C{row}"].font = Font(size=9, italic=True)

        row += 2
        sheet[f"A{row}"] = "____________________________"
        sheet[f"C{row}"] = "____________________________"

        row += 1
        sheet[f"A{row}"] = "Date"
        sheet[f"A{row}"].font = Font(size=9, italic=True)
        sheet[f"C{row}"] = "Date"
        sheet[f"C{row}"].font = Font(size=9, italic=True)

        # Footer
        row += 3
        sheet[f"A{row}"] = "Generated by Painting.AI - Professional Estimating Software"
        sheet[f"A{row}"].font = Font(size=8, italic=True, color="666666")
        sheet.merge_cells(f'A{row}:E{row}')
        sheet[f"A{row}"].alignment = Alignment(horizontal='center')

    def _create_takeoff_sheet(self, workbook, project: Dict, rooms: List[Dict]):
        """Create detailed takeoff sheet"""

        sheet = workbook.create_sheet("Detailed Takeoff")

        # Header
        sheet["A1"] = "DETAILED PAINTING TAKEOFF"
        sheet["A1"].font = Font(size=14, bold=True)

        row = 3
        headers = ["Room", "Surface", "Description", "Quantity", "Unit", "Gallons", "Hours", "Cost"]

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row += 1

        # Add line items for each room
        for room in rooms:
            room_name = room.get("name", "")
            surfaces = room.get("surfaces", {})

            for surface_name, surface_data in surfaces.items():
                area = surface_data.get("area", 0)

                # Primer line
                sheet.cell(row=row, column=1).value = room_name
                sheet.cell(row=row, column=2).value = surface_name.title()
                sheet.cell(row=row, column=3).value = "Primer - 1 coat"
                sheet.cell(row=row, column=4).value = round(area, 1)
                sheet.cell(row=row, column=5).value = "sqft"
                sheet.cell(row=row, column=6).value = round(area / 400, 2)
                sheet.cell(row=row, column=7).value = round(area / 300 * 1.2, 2)
                sheet.cell(row=row, column=8).value = ""  # Cost calculated later
                row += 1

                # Finish coat line
                sheet.cell(row=row, column=1).value = room_name
                sheet.cell(row=row, column=2).value = surface_name.title()
                sheet.cell(row=row, column=3).value = "Finish - 2 coats"
                sheet.cell(row=row, column=4).value = round(area * 2, 1)
                sheet.cell(row=row, column=5).value = "sqft"
                sheet.cell(row=row, column=6).value = round(area * 2 / 400, 2)
                sheet.cell(row=row, column=7).value = round(area * 2 / 300 * 1.2, 2)
                sheet.cell(row=row, column=8).value = ""
                row += 1

        # Auto-size columns
        for col in range(1, 9):
            sheet.column_dimensions[get_column_letter(col)].width = 15

    def _create_room_breakdown_sheet(self, workbook, project: Dict, rooms: List[Dict]):
        """Create room-by-room breakdown sheet"""

        sheet = workbook.create_sheet("Room Breakdown")

        row = 1

        for room in rooms:
            # Room header
            sheet[f"A{row}"] = room.get("name", "").upper()
            sheet[f"A{row}"].font = Font(size=12, bold=True)
            row += 1

            # Dimensions
            dims = room.get("dimensions", {})
            if dims:
                sheet[f"A{row}"] = "Dimensions:"
                sheet[f"B{row}"] = f"{dims.get('length', 0)}' L × {dims.get('width', 0)}' W × {dims.get('height', 9)}' H"
                row += 1

            # Surfaces
            sheet[f"A{row}"] = "Surface"
            sheet[f"B{row}"] = "Area (sqft)"
            sheet[f"A{row}"].font = Font(bold=True)
            sheet[f"B{row}"].font = Font(bold=True)
            row += 1

            surfaces = room.get("surfaces", {})
            for surface_name, surface_data in surfaces.items():
                sheet[f"A{row}"] = surface_name.title()
                sheet[f"B{row}"] = round(surface_data.get("area", 0), 1)
                row += 1

            # Total
            sheet[f"A{row}"] = "TOTAL PAINTABLE AREA"
            sheet[f"B{row}"] = round(room.get("total_area", 0), 1)
            sheet[f"A{row}"].font = Font(bold=True)
            sheet[f"B{row}"].font = Font(bold=True)
            row += 2

        # Auto-size columns
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15

    def generate_pdf(self, project: Dict, rooms: List[Dict], output_path: Path):
        """Generate PDF proposal"""

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1f4788"),
            spaceAfter=30,
            alignment=1  # Center
        )

        story.append(Paragraph("PAINTING PROPOSAL", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Project info
        info_data = [
            ["Project:", project.get("name", "")],
            ["Customer:", project.get("customer", "")],
            ["Date:", datetime.now().strftime("%B %d, %Y")],
        ]

        info_table = Table(info_data, colWidths=[1.5 * inch, 4 * inch])
        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))

        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        # Summary
        story.append(Paragraph("PROJECT SUMMARY", styles["Heading2"]))

        summary_data = [
            ["Total Rooms:", f"{project.get('total_rooms', 0)}"],
            ["Total Paintable Area:", f"{project.get('total_sqft', 0):,.0f} sqft"],
            ["Total Paint Needed:", f"{project.get('total_gallons', 0):,.0f} gallons"],
            ["Total Labor Hours:", f"{project.get('total_labor_hours', 0):,.1f} hours"],
        ]

        summary_table = Table(summary_data, colWidths=[2.5 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f0f0")),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # Room list
        story.append(Paragraph("ROOM BREAKDOWN", styles["Heading2"]))

        room_data = [["Room Name", "Total Area", "Walls", "Ceiling", "Trim"]]

        for room in rooms:
            surfaces = room.get("surfaces", {})
            room_data.append([
                room.get("name", ""),
                f"{room.get('total_area', 0):,.0f} sqft",
                f"{surfaces.get('walls', {}).get('area', 0):,.0f}",
                f"{surfaces.get('ceiling', {}).get('area', 0):,.0f}",
                f"{surfaces.get('trim', {}).get('area', 0):,.0f}",
            ])

        room_table = Table(room_data, colWidths=[2 * inch, 1.2 * inch, 1 * inch, 1 * inch, 1 * inch])
        room_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9f9f9")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))

        story.append(room_table)
        story.append(Spacer(1, 0.5 * inch))

        # Total cost
        cost_data = [[
            "TOTAL ESTIMATED COST:",
            f"${project.get('estimated_cost', 0):,.2f}"
        ]]

        cost_table = Table(cost_data, colWidths=[4 * inch, 2 * inch])
        cost_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 14),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1f4788")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.whitesmoke),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))

        story.append(cost_table)

        # Build PDF
        doc.build(story)


if __name__ == "__main__":
    # Test export generation
    from pathlib import Path

    test_project = {
        "name": "Office Renovation",
        "customer": "ABC Construction",
        "total_rooms": 3,
        "total_sqft": 2400,
        "total_gallons": 18,
        "total_labor_hours": 48,
        "estimated_cost": 3500
    }

    test_rooms = [
        {
            "name": "Conference Room",
            "total_area": 900,
            "dimensions": {"length": 20, "width": 15, "height": 9},
            "surfaces": {
                "walls": {"area": 600},
                "ceiling": {"area": 300},
                "trim": {"area": 50}
            }
        },
        {
            "name": "Office 1",
            "total_area": 750,
            "dimensions": {"length": 15, "width": 12, "height": 9},
            "surfaces": {
                "walls": {"area": 486},
                "ceiling": {"area": 180},
                "trim": {"area": 40}
            }
        }
    ]

    exporter = ExportGenerator()

    excel_path = Path("test_export.xlsx")
    pdf_path = Path("test_proposal.pdf")

    exporter.generate_excel(test_project, test_rooms, excel_path)
    print(f"✓ Excel generated: {excel_path}")

    exporter.generate_pdf(test_project, test_rooms, pdf_path)
    print(f"✓ PDF generated: {pdf_path}")
