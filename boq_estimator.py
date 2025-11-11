"""
BOQ Estimator - Main Integration Module
Complete workflow from construction drawings to validated BOQ

This is the main entry point that orchestrates the entire estimation process:
1. Extract data from drawings using AI
2. Map to BOQ structure
3. Calculate rates
4. Validate against CPWD standards
5. Generate final BOQ document
"""

import json
import logging
from pathlib import Path
from typing import List, Dict, Optional, Union
from decimal import Decimal
from datetime import datetime

from boq_schema import BOQDocument, ContractDetails, BOQSection
from drawing_extractor import DrawingExtractor, DrawingExtractionResult
from boq_mapper import BOQMapper, BOQAutoNumberer, merge_similar_items
from boq_validator import CPWDValidator, ValidationResult
from boq_calculator import get_calculator

logger = logging.getLogger(__name__)


class BOQEstimator:
    """
    Complete BOQ estimation system
    Orchestrates the entire workflow from drawings to final BOQ
    """

    def __init__(
        self,
        anthropic_api_key: str,
        dsr_rates: Optional[Dict] = None,
        enable_logging: bool = True
    ):
        """
        Initialize BOQ Estimator

        Args:
            anthropic_api_key: Anthropic API key for Claude Vision
            dsr_rates: Dictionary of DSR rates (optional)
            enable_logging: Enable detailed logging
        """
        self.drawing_extractor = DrawingExtractor(anthropic_api_key)
        self.mapper = BOQMapper(dsr_rates)
        self.validator = CPWDValidator()
        self.dsr_rates = dsr_rates or self._load_default_rates()

        if enable_logging:
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )

    def estimate_from_drawings(
        self,
        drawing_paths: List[Union[str, Path]],
        contract_details: ContractDetails,
        output_dir: Optional[Union[str, Path]] = None,
        merge_duplicates: bool = True,
        validate: bool = True
    ) -> BOQDocument:
        """
        Complete estimation workflow from drawings to BOQ

        Args:
            drawing_paths: List of drawing file paths (PDF or images)
            contract_details: Contract metadata
            output_dir: Optional directory for intermediate outputs
            merge_duplicates: Merge similar items to reduce duplication
            validate: Perform CPWD validation

        Returns:
            Complete BOQ document
        """
        logger.info("=" * 80)
        logger.info("BOQ ESTIMATION WORKFLOW STARTED")
        logger.info("=" * 80)

        # Step 1: Extract from drawings
        logger.info("\n[1/5] Extracting data from drawings...")
        extraction_results = self.drawing_extractor.batch_process(
            drawing_paths,
            output_dir=Path(output_dir) / "extractions" if output_dir else None
        )
        logger.info(f"✓ Extracted data from {len(extraction_results)} drawing pages")

        # Step 2: Map to BOQ structure
        logger.info("\n[2/5] Mapping extracted data to BOQ structure...")
        sections = self.mapper.map_extraction_to_boq(
            extraction_results,
            project_name=contract_details.contract_name
        )
        logger.info(f"✓ Created {len(sections)} BOQ sections")

        # Step 3: Merge duplicates if requested
        if merge_duplicates:
            logger.info("\n[3/5] Merging duplicate items...")
            for section in sections:
                section.items = merge_similar_items(section.items)
            logger.info("✓ Merged duplicate items")
        else:
            logger.info("\n[3/5] Skipping duplicate merge")

        # Step 4: Renumber items
        logger.info("\n[4/5] Renumbering BOQ items...")
        sections = BOQAutoNumberer.renumber_sections(sections)
        logger.info("✓ Renumbered all items")

        # Step 5: Create BOQ document
        logger.info("\n[5/5] Creating BOQ document...")
        boq_doc = BOQDocument(
            boq_id=f"BOQ-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
            contract=contract_details,
            sections=sections
        )
        logger.info(f"✓ BOQ document created with total amount: ₹{boq_doc.total_amount:,.2f}")

        # Validation
        if validate:
            logger.info("\nValidating BOQ against CPWD standards...")
            is_valid, validation_results = self.validator.validate_document(boq_doc)

            if is_valid:
                logger.info("✓ BOQ validation passed!")
            else:
                logger.warning("⚠ BOQ has validation issues")
                self.validator.print_validation_report()

        # Save outputs
        if output_dir:
            self._save_outputs(boq_doc, output_dir, extraction_results)

        logger.info("\n" + "=" * 80)
        logger.info("BOQ ESTIMATION WORKFLOW COMPLETED")
        logger.info("=" * 80)

        return boq_doc

    def estimate_from_extraction_results(
        self,
        extraction_results: List[DrawingExtractionResult],
        contract_details: ContractDetails,
        validate: bool = True
    ) -> BOQDocument:
        """
        Create BOQ from pre-extracted data

        Use this if you already have extraction results and want to skip
        the drawing extraction step.

        Args:
            extraction_results: Pre-extracted drawing data
            contract_details: Contract metadata
            validate: Perform validation

        Returns:
            BOQ document
        """
        sections = self.mapper.map_extraction_to_boq(
            extraction_results,
            project_name=contract_details.contract_name
        )

        sections = BOQAutoNumberer.renumber_sections(sections)

        boq_doc = BOQDocument(
            boq_id=f"BOQ-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
            contract=contract_details,
            sections=sections
        )

        if validate:
            self.validator.validate_document(boq_doc)

        return boq_doc

    def _save_outputs(
        self,
        boq_doc: BOQDocument,
        output_dir: Union[str, Path],
        extraction_results: Optional[List[DrawingExtractionResult]] = None
    ):
        """Save BOQ and related outputs to directory"""
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True, parents=True)

        # Save BOQ as JSON
        boq_json_path = output_dir / "boq_document.json"
        with open(boq_json_path, 'w') as f:
            json.dump(
                json.loads(boq_doc.json()),
                f,
                indent=2,
                default=str
            )
        logger.info(f"Saved BOQ JSON: {boq_json_path}")

        # Save summary
        summary_path = output_dir / "boq_summary.json"
        summary = {
            'boq_id': boq_doc.boq_id,
            'contract_name': boq_doc.contract.contract_name,
            'total_amount': float(boq_doc.total_amount),
            'total_sections': len(boq_doc.sections),
            'total_items': sum(len(s.items) for s in boq_doc.sections),
            'category_wise_summary': {
                str(k): float(v)
                for k, v in boq_doc.get_summary_by_category().items()
            },
            'created_date': boq_doc.created_date.isoformat()
        }
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2)
        logger.info(f"Saved summary: {summary_path}")

        # Save validation report
        if self.validator.validation_results:
            validation_path = output_dir / "validation_report.json"
            validation_data = {
                'summary': self.validator.get_validation_summary(),
                'issues': [r.to_dict() for r in self.validator.validation_results]
            }
            with open(validation_path, 'w') as f:
                json.dump(validation_data, f, indent=2)
            logger.info(f"Saved validation report: {validation_path}")

    def _load_default_rates(self) -> Dict:
        """Load default DSR rates"""
        # This would typically load from a database or rate file
        # Here we provide sample rates for demonstration
        return {
            # Labour rates (per day)
            'LABOUR_UNSKILLED': Decimal("600"),
            'LABOUR_SKILLED': Decimal("800"),
            'LABOUR_MASON': Decimal("850"),
            'LABOUR_HELPER': Decimal("600"),
            'LABOUR_CARPENTER': Decimal("900"),
            'LABOUR_BARBENDER': Decimal("850"),

            # Material rates
            'CEMENT_OPC43': Decimal("7.50"),  # per kg
            'SAND': Decimal("1200"),  # per cum
            'AGGREGATE_20MM': Decimal("1400"),  # per cum
            'WATER': Decimal("0.05"),  # per litre
            'BRICK_FIRST_CLASS': Decimal("6.00"),  # per piece
            'STEEL_FE500': Decimal("65"),  # per kg
            'BINDING_WIRE': Decimal("80"),  # per kg
            'PLYWOOD_12MM': Decimal("1800"),  # per sqm
            'TIMBER_BATTEN': Decimal("60"),  # per rmt
        }

    def export_to_excel(self, boq_doc: BOQDocument, output_path: Union[str, Path]):
        """
        Export BOQ to Excel format (requires openpyxl)

        Args:
            boq_doc: BOQ document
            output_path: Path for Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOQ"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Title
        ws['A1'] = boq_doc.contract.contract_name
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"BOQ ID: {boq_doc.boq_id}"
        ws['A3'] = f"Date: {boq_doc.created_date.strftime('%d-%m-%Y')}"

        # Column headers
        row = 5
        headers = ['Item No', 'Description', 'Specification', 'Unit', 'Quantity', 'Rate', 'Amount']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # BOQ items
        for section in boq_doc.sections:
            # Section header
            ws.cell(row=row, column=1).value = f"Section {section.section_no}: {section.title}"
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # Items
            for item in section.items:
                ws.cell(row=row, column=1).value = item.item_no
                ws.cell(row=row, column=2).value = item.description
                ws.cell(row=row, column=3).value = item.specification
                ws.cell(row=row, column=4).value = item.unit.value
                ws.cell(row=row, column=5).value = float(item.quantity)
                ws.cell(row=row, column=6).value = float(item.unit_rate)
                ws.cell(row=row, column=7).value = float(item.total_amount)
                row += 1

            # Section total
            ws.cell(row=row, column=6).value = "Section Total:"
            ws.cell(row=row, column=6).font = Font(bold=True)
            ws.cell(row=row, column=7).value = float(section.total_amount)
            ws.cell(row=row, column=7).font = Font(bold=True)
            row += 2

        # Grand total
        ws.cell(row=row, column=6).value = "GRAND TOTAL:"
        ws.cell(row=row, column=6).font = Font(bold=True, size=12)
        ws.cell(row=row, column=7).value = float(boq_doc.total_amount)
        ws.cell(row=row, column=7).font = Font(bold=True, size=12)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        logger.info(f"Exported BOQ to Excel: {output_path}")


# Convenience function for quick estimation
def quick_estimate(
    drawing_paths: List[str],
    project_name: str,
    location: str,
    anthropic_api_key: str,
    output_dir: str = "./boq_output"
) -> BOQDocument:
    """
    Quick estimation function with minimal configuration

    Args:
        drawing_paths: List of drawing file paths
        project_name: Project name
        location: Project location
        anthropic_api_key: Anthropic API key
        output_dir: Output directory

    Returns:
        BOQ document
    """
    contract = ContractDetails(
        contract_name=project_name,
        location=location,
        client="Client Name",
        department="CPWD"
    )

    estimator = BOQEstimator(anthropic_api_key)

    return estimator.estimate_from_drawings(
        drawing_paths=drawing_paths,
        contract_details=contract,
        output_dir=output_dir
    )


if __name__ == "__main__":
    # Example usage
    print("BOQ Estimator - Production Ready")
    print("=" * 80)
    print("\nThis module provides complete BOQ estimation from construction drawings.")
    print("\nBasic usage:")
    print("""
    from boq_estimator import BOQEstimator, ContractDetails

    # Initialize estimator
    estimator = BOQEstimator(anthropic_api_key="your-api-key")

    # Define contract
    contract = ContractDetails(
        contract_name="Office Building Construction",
        location="New Delhi",
        client="Government Department",
        department="CPWD"
    )

    # Process drawings and generate BOQ
    boq = estimator.estimate_from_drawings(
        drawing_paths=["drawing1.pdf", "drawing2.pdf"],
        contract_details=contract,
        output_dir="./output"
    )

    # Export to Excel
    estimator.export_to_excel(boq, "boq_final.xlsx")
    """)
